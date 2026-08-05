#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera planilha integrada do IBrX100 - versao que evita conflito
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import time

# Carrega scoring
with open('D:\\Meus APP\\ibrx100_system\\output\\data\\scores.json', 'r', encoding='utf-8') as f:
    scores = json.load(f)

top10 = scores[:10]

# Cria workbook
wb = openpyxl.Workbook()

# Estilos
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
score_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================
# SHEET 1: CARTEIRA INTEGRADA
# ============================================================
ws_carteira = wb.active
ws_carteira.title = "Carteira IBrX100"

headers_carteira = [
    "Rank", "Ticker", "Nome", "Score", "Fundamental", "Valuation", 
    "Momentum", "Dividendos", "Data Entrada", "Preco Entrada", 
    "Quantidade", "Custo Total", "Preco Atual", "Valor Atual", 
    "Lucro/Prejuizo", "Retorno %"
]

for col, header in enumerate(headers_carteira, 1):
    cell = ws_carteira.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

nomes = {
    "FLRY3.SA": "Fleury",
    "ABEV3.SA": "Ambev",
    "PETR3.SA": "Petrobras PN",
    "PETR4.SA": "Petrobras ON",
    "HYPE3.SA": "Hypera",
    "EZTC3.SA": "Eztec",
    "MULT3.SA": "Multiplan",
    "B3SA3.SA": "B3",
    "TOTS3.SA": "TOTVS",
    "LREN3.SA": "Lojas Renner"
}

for i, stock in enumerate(top10):
    row = i + 2
    ticker = stock['ticker']
    ticker_b3 = ticker.replace('.SA', '')
    
    ws_carteira.cell(row=row, column=1, value=i+1).border = thin_border
    ws_carteira.cell(row=row, column=2, value=ticker_b3).border = thin_border
    ws_carteira.cell(row=row, column=3, value=nomes.get(ticker, '')).border = thin_border
    
    cell_score = ws_carteira.cell(row=row, column=4, value=stock['score_composto'])
    cell_score.fill = score_fill
    cell_score.border = thin_border
    
    ws_carteira.cell(row=row, column=5, value=stock['score_fundamental']).border = thin_border
    ws_carteira.cell(row=row, column=6, value=stock['score_valuation']).border = thin_border
    ws_carteira.cell(row=row, column=7, value=stock['score_momentum']).border = thin_border
    ws_carteira.cell(row=row, column=8, value=stock['score_dividendos']).border = thin_border
    
    ws_carteira.cell(row=row, column=9, value='=TODAY()').border = thin_border
    ws_carteira.cell(row=row, column=10, value=f'=IFERROR(GOOGLEFINANCE("BVMF:{ticker_b3}"),{stock["preco_atual"]})').border = thin_border
    ws_carteira.cell(row=row, column=11, value=0).border = thin_border
    ws_carteira.cell(row=row, column=12, value=f'=J{row}*K{row}').border = thin_border
    ws_carteira.cell(row=row, column=13, value=f'=IFERROR(GOOGLEFINANCE("BVMF:{ticker_b3}","price"),{stock["preco_atual"]})').border = thin_border
    ws_carteira.cell(row=row, column=14, value=f'=K{row}*M{row}').border = thin_border
    ws_carteira.cell(row=row, column=15, value=f'=N{row}-L{row}').border = thin_border
    ws_carteira.cell(row=row, column=16, value=f'=IFERROR(O{row}/L{row},0)').border = thin_border

col_widths = [6, 10, 15, 7, 11, 10, 9, 10, 12, 12, 11, 12, 12, 12, 14, 10]
for i, width in enumerate(col_widths, 1):
    ws_carteira.column_dimensions[get_column_letter(i)].width = width

# ============================================================
# SHEET 2: RANKING COMPLETO
# ============================================================
ws_ranking = wb.create_sheet("Ranking Completo")

headers_ranking = [
    "Rank", "Ticker", "Score", "Fundamental", "Valuation", 
    "Momentum", "Dividendos", "Preco Atual", "MM50", "MM200", "Euforia"
]

for col, header in enumerate(headers_ranking, 1):
    cell = ws_ranking.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

for i, stock in enumerate(scores):
    row = i + 2
    ticker = stock['ticker'].replace('.SA', '')
    
    ws_ranking.cell(row=row, column=1, value=i+1).border = thin_border
    ws_ranking.cell(row=row, column=2, value=ticker).border = thin_border
    ws_ranking.cell(row=row, column=3, value=stock['score_composto']).border = thin_border
    ws_ranking.cell(row=row, column=4, value=stock['score_fundamental']).border = thin_border
    ws_ranking.cell(row=row, column=5, value=stock['score_valuation']).border = thin_border
    ws_ranking.cell(row=row, column=6, value=stock['score_momentum']).border = thin_border
    ws_ranking.cell(row=row, column=7, value=stock['score_dividendos']).border = thin_border
    ws_ranking.cell(row=row, column=8, value=stock['preco_atual']).border = thin_border
    ws_ranking.cell(row=row, column=9, value=stock['mm50']).border = thin_border
    ws_ranking.cell(row=row, column=10, value=stock['mm200']).border = thin_border
    ws_ranking.cell(row=row, column=11, value=stock['euforia']).border = thin_border

col_widths_ranking = [6, 10, 7, 11, 10, 9, 10, 12, 12, 12, 8]
for i, width in enumerate(col_widths_ranking, 1):
    ws_ranking.column_dimensions[get_column_letter(i)].width = width

# ============================================================
# SHEET 3: INSTRUCOES
# ============================================================
ws_instrucoes = wb.create_sheet("Instrucoes")

instrucoes = [
    ("INSTRUCOES DE USO", ""),
    ("", ""),
    ("1. CARTERA IBrX100", "Top 10 acoes do scoring automatico"),
    ("2. RANKING COMPLETO", "Todos os 20+ acoes com scores"),
    ("", ""),
    ("COMO USAR:", ""),
    ("1. Preencha a coluna 'Quantidade' na aba Carteira", ""),
    ("2. As formulas GOOGLEFINANCE atualizam automaticamente", ""),
    ("3. Para atualizar manualmente: Dados > Atualizar tudo", ""),
    ("", ""),
    ("FORMULAS:", ""),
    ("Preco Entrada: =GOOGLEFINANCE('BVMF:TICKER')", ""),
    ("Preco Atual: =GOOGLEFINANCE('BVMF:TICKER','price')", ""),
    ("Custo Total: =Preco_Entrada * Quantidade", ""),
    ("Valor Atual: =Preco_Atual * Quantidade", ""),
    ("Lucro/Prejuizo: =Valor_Atual - Custo_Total", ""),
    ("Retorno %: =Lucro/Prejuizo / Custo_Total", ""),
    ("", ""),
    ("ATUALIZACAO:", ""),
    ("Execute: python gerar_carteira.py", ""),
    ("", ""),
    ("DATA:", datetime.now().strftime("%d/%m/%Y %H:%M"))
]

for row_idx, (col1, col2) in enumerate(instrucoes, 1):
    ws_instrucoes.cell(row=row_idx, column=1, value=col1)
    ws_instrucoes.cell(row=row_idx, column=2, value=col2)

ws_instrucoes.column_dimensions['A'].width = 50
ws_instrucoes.column_dimensions['B'].width = 50

# Salva com nome temporario
temp_path = 'G:\\Meu Drive\\Pessoal\\Meus APP\\Investimentos\\_temp_carteira.xlsx'
final_path = 'G:\\Meu Drive\\Pessoal\\Meus APP\\Investimentos\\Carteira IBrX100 integrada.xlsx'

# Remove arquivo temporario anterior se existir
if os.path.exists(temp_path):
    os.remove(temp_path)

# Salva como temporario
wb.save(temp_path)
print(f"Salvo como temporario: {temp_path}")

# Espera um pouco
time.sleep(2)

# Remove o antigo se existir
if os.path.exists(final_path):
    try:
        os.remove(final_path)
        print(f"Removido: {final_path}")
    except:
        print("Nao foi possivel remover o arquivo antigo")
        print("Por favor, feche o Excel e execute novamente")
        
# Renomeia
try:
    os.rename(temp_path, final_path)
    print(f"Renomeado para: {final_path}")
except:
    print(f"Arquivo temporario salvo em: {temp_path}")
    print("Por favor, renomeie manualmente para 'Carteira IBrX100 integrada.xlsx'")

print("\nTop 10 incluidos:")
for i, stock in enumerate(top10):
    ticker = stock['ticker'].replace('.SA', '')
    print(f"  {i+1}. {ticker} - Score: {stock['score_composto']}")