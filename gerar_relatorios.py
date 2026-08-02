#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Gera relatórios PDF e XLSX"""

import json
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Carrega dados
with open('data/scores.json', 'r', encoding='utf-8') as f:
    dados = json.load(f)

with open('data/fundamentais.json', 'r', encoding='utf-8') as f:
    fundamentais = json.load(f)

data_hoje = datetime.now().strftime('%d/%m/%Y')

# ============================================================
# RELATÓRIO PDF
# ============================================================
def gerar_pdf():
    doc = SimpleDocTemplate(
        "output/relatorio_trading.pdf",
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30
    )
    
    elements = []
    
    # Título
    elements.append(Paragraph("Relatório de Trading - B3", title_style))
    elements.append(Paragraph(f"Data: {data_hoje}", styles['Normal']))
    elements.append(Spacer(1, 1*cm))
    
    # Ranking
    elements.append(Paragraph("Ranking das Ações", styles['Heading2']))
    
    ranking_data = [['#', 'Ação', 'Score', 'Fund', 'Mom', 'Val', 'Div', 'Preço', 'DY']]
    for i, item in enumerate(dados, 1):
        ranking_data.append([
            str(i),
            item['ticker'],
            str(item['score_composto']),
            str(item['score_fundamental']),
            str(item['score_momentum']),
            str(item['score_valuation']),
            str(item['score_dividendos']),
            f"R$ {item['preco']:.2f}" if item['preco'] else "-",
            f"{item['dy']:.1f}%" if item['dy'] else "-"
        ])
    
    ranking_table = Table(ranking_data, colWidths=[1*cm, 2*cm, 1.5*cm, 1.5*cm, 1.5*cm, 1.5*cm, 1.5*cm, 2.5*cm, 2*cm])
    ranking_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(ranking_table)
    elements.append(Spacer(1, 1*cm))
    
    # Detalhes por ação
    elements.append(Paragraph("Detalhes por Ação", styles['Heading2']))
    
    for item in dados:
        ticker = item['ticker']
        fund = fundamentais.get(ticker + '.SA', {})
        
        elements.append(Paragraph(f"<b>{ticker}</b> - {item['nome']}", styles['Heading3']))
        
        detalhes = [
            ['Indicador', 'Valor'],
            ['Preço Atual', f"R$ {item['preco']:.2f}" if item['preco'] else "-"],
            ['P/L', f"{item['pl']:.1f}" if item['pl'] else "-"],
            ['P/VP', f"{item['pvp']:.2f}" if item['pvp'] else "-"],
            ['ROE', f"{item['roe']:.1%}" if item['roe'] else "-"],
            ['DY', f"{item['dy']:.1f}%" if item['dy'] else "-"],
            ['Setor', item['setor'] if item['setor'] else "-"]
        ]
        
        detalhes_table = Table(detalhes, colWidths=[4*cm, 6*cm])
        detalhes_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(detalhes_table)
        elements.append(Spacer(1, 0.5*cm))
    
    # Gestão de Risco
    elements.append(Paragraph("Gestão de Risco", styles['Heading2']))
    
    risco_data = [
        ['Parâmetro', 'Valor'],
        ['Stop Loss', '10%'],
        ['Trailing Stop', '8%'],
        ['Máximo por Ação', '5%'],
        ['Máximo Posições', '5'],
        ['Capital Inicial', 'R$ 10.000']
    ]
    
    risco_table = Table(risco_data, colWidths=[6*cm, 6*cm])
    risco_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(risco_table)
    
    doc.build(elements)
    print("PDF gerado: output/relatorio_trading.pdf")

# ============================================================
# RELATÓRIO XLSX
# ============================================================
def gerar_xlsx():
    wb = openpyxl.Workbook()
    
    # Aba Ranking
    ws = wb.active
    ws.title = "Ranking"
    
    # Cabeçalho
    headers = ['#', 'Ação', 'Nome', 'Score', 'Fundamental', 'Momentum', 'Valuation', 'Dividendos', 'Preço', 'P/L', 'P/VP', 'ROE', 'DY', 'Setor']
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Dados
    for row, item in enumerate(dados, 2):
        ws.cell(row=row, column=1, value=row-1).border = thin_border
        ws.cell(row=row, column=2, value=item['ticker']).border = thin_border
        ws.cell(row=row, column=3, value=item['nome']).border = thin_border
        ws.cell(row=row, column=4, value=item['score_composto']).border = thin_border
        ws.cell(row=row, column=5, value=item['score_fundamental']).border = thin_border
        ws.cell(row=row, column=6, value=item['score_momentum']).border = thin_border
        ws.cell(row=row, column=7, value=item['score_valuation']).border = thin_border
        ws.cell(row=row, column=8, value=item['score_dividendos']).border = thin_border
        ws.cell(row=row, column=9, value=item['preco']).border = thin_border
        ws.cell(row=row, column=10, value=item['pl']).border = thin_border
        ws.cell(row=row, column=11, value=item['pvp']).border = thin_border
        ws.cell(row=row, column=12, value=item['roe']).border = thin_border
        ws.cell(row=row, column=13, value=item['dy']).border = thin_border
        ws.cell(row=row, column=14, value=item['setor']).border = thin_border
    
    # Ajusta largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2
    
    # Aba Gestão de Risco
    ws_risco = wb.create_sheet("Gestão de Risco")
    ws_risco['A1'] = 'Parâmetro'
    ws_risco['B1'] = 'Valor'
    ws_risco['A1'].font = header_font
    ws_risco['A1'].fill = header_fill
    ws_risco['B1'].font = header_font
    ws_risco['B1'].fill = header_fill
    
    risco = [
        ('Stop Loss', '10%'),
        ('Trailing Stop', '8%'),
        ('Máximo por Ação', '5%'),
        ('Máximo Posições', '5'),
        ('Capital Inicial', 'R$ 10.000')
    ]
    
    for i, (param, valor) in enumerate(risco, 2):
        ws_risco.cell(row=i, column=1, value=param)
        ws_risco.cell(row=i, column=2, value=valor)
    
    ws_risco.column_dimensions['A'].width = 20
    ws_risco.column_dimensions['B'].width = 15
    
    # Salva
    wb.save("output/relatorio_trading.xlsx")
    print("XLSX gerado: output/relatorio_trading.xlsx")

# ============================================================
if __name__ == '__main__':
    gerar_pdf()
    gerar_xlsx()
